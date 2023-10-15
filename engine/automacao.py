"""
Este módulo é voltado para guardar o código do Web Scraping. 

Neste módulo é usando a classe Navegador, que herda as funcionalidades do Selenium.

O objetivo deste código é coletar os dados de figuras públicas através da rede social do Twitter.

Documentação Adicional: https://peps.python.org/pep-0008/ | https://selenium-python.readthedocs.io/index.html

Licença:
Este código está sujeito às políticas e regulamentos internos.

© 2023, Natan.
"""

# Importações do Python
import requests
import datetime
from openpyxl import Workbook, load_workbook
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# Importações Internas
from engine.navegador import Navegador
from utilitarios.utilitarios import *

class Automacao:
    def raspar_Twitter(self):
        """Função definida para raspar os dados de um perfil do Twitter."""

        # Salvando o input do usuário
        url = input("\033[34mInsira o nome do usuário ou cole o link do perfil: \033[0m")
        usuario_pesquisa = url.split("/")[-1].split("?")[0]

        # Verifica se deseja período
        filtrar_periodo = False
        inp = input("\033[34mDeseja informar período? [s/n]: \033[0m")
        if  inp.upper().strip() == "S":
            periodo_inicio = input("\033[34mInforme o início do período que deseja filtrar - [DD/MM/YYYY]: \033[0m")
            periodo_final = input("\033[34mInforme o final do período que deseja filtrar - [DD/MM/YYYY]: \033[0m")
            filtrar_periodo = True
        sleep(0.5)
        self.carregar()
        os.system('cls')
        self.carregar_iniciando()

        # Instanciando o Navegador
        self.navegador = Navegador(salvar_cache=PATH['arquivo_cache'], tela_cheia=True)

        # Verficia de o Input foi uma URL ou o nome do Usuário
        if self.e_url(url):
            self.navegador.navegar(url)
        else: 
            self.navegador.navegar(LINKS['base_twitter'] + usuario_pesquisa)
        sleep(3)

        # Verifica login
        if len(self.navegador.obter_elementos('XPATH', XPATHS['login'])) > 0:
            print("Deve fazer login..")
            self.login(CREDENCIAIS['usuario_twitter'], CREDENCIAIS['senha_twitter'])
            if self.e_url(url):
                self.navegador.navegar(url)
            else:
                self.navegador.navegar(LINKS['base_twitter'] + usuario_pesquisa)
            sleep(3)

        # Esperar carregar página
        self.navegador.esperar('XPATH', XPATHS['posts'])
        sleep(1)
        print("Coletando dados do perfil:", usuario_pesquisa)

        i = 0
        parar = 0
        qtde_arquivos = 0
        qtde_processados = 0
        processados = []
        # Enquanto tiver post
        while 1:
            nome_arquivos_imagens = ""
            nome_arquivos_video = ""
            texto_postagem = ""
            data_atual = datetime.datetime.now().strftime("%d.%m.%Y")
            posts = self.navegador.obter_elementos('XPATH', XPATHS['posts'])

            for post in posts:
                i += 1

                # Verifica se o post já foi processado
                post_id = post.get_attribute('aria-labelledby')
                if post_id in processados:
                    print('ja foi processado')
                    qtde_processados += 1
                    i -= 1
                    continue
                
                # Salva os dados
                self.navegador.esperar('xpath', XPATHS['posts'])
                sleep(0.3)
                try:
                    texto_postagem = post.find_element(By.XPATH, XPATHS['texto']).text.replace("\n", " ")
                except:
                    texto_postagem = ""
                url_postagem = post.find_element(By.XPATH, XPATHS['link_data']).get_attribute("href")

                # Verifica se tem Min ou Hora no texto para saber se o post é do dia atual
                data = post.find_element(By.XPATH, XPATHS['link_data']).text
                if data[-1].strip().lower() == 'm' or data[-1].strip().lower() == 'h':
                    data_postagem = datetime.datetime.now().strftime("%d/%m/%Y")
                else:
                    dia = data.split(" ")[1].split(",")[0].zfill(2)
                    mes = str(self.converte_nome_mes(data.split(" ")[0])).zfill(2)
                    if data.split(" ")[-1].isdigit() and len(data.split(" ")[-1]) == 4:
                        ano = data.split(" ")[-1]
                    else:
                        ano = datetime.datetime.now().year
                    data_postagem = f"{dia}/{mes}/{ano}"

                # Verifica se a data está dentro do período
                if filtrar_periodo:
                    data_inicio = datetime.datetime.strptime(periodo_inicio, "%d/%m/%Y")
                    data_verificar = datetime.datetime.strptime(data_postagem, "%d/%m/%Y")
                    data_final = datetime.datetime.strptime(periodo_final, "%d/%m/%Y")

                    # Se não tiver dentro do períodom, encerra. 
                    if not data_inicio <= data_verificar <= data_final:
                        
                        # Verifica se o post não é um post fixado
                        if len(post.find_elements(By.XPATH, XPATHS['fixado'])) == 0:
                            parar = 1
                            break

                # Verifica se o arquivo de OUTPUT está criado
                nome_arquivo = f"\{usuario_pesquisa}\{data_atual}"
                if not os.path.exists(PATH['saida'] + f'\{usuario_pesquisa}'):
                    os.mkdir(PATH['saida'] + f'\{usuario_pesquisa}')
                if not os.path.exists(PATH['saida'] + f'\{nome_arquivo}'):
                    os.mkdir(PATH['saida'] + f'\{nome_arquivo}')
                
                # Verifica se tem imagem
                if len(post.find_elements(By.CSS_SELECTOR, 'img[alt="Image"]')) > 0:
                    imagens = post.find_elements(By.CSS_SELECTOR, 'img[alt="Image"]')
                    print(f"Salvando Imagem do post #{i}.")
                    for img in imagens:
                        imagem_url = img.get_attribute("src")
                        hora = str(datetime.datetime.now().hour) + str(datetime.datetime.now().minute) + str(datetime.datetime.now().second) + str(datetime.datetime.now().microsecond)
                        if not os.path.exists(PATH['saida'] + nome_arquivo + f"\img"):
                            os.makedirs(PATH['saida'] + nome_arquivo + f"\img")
                        self.baixar_conteudo(imagem_url, PATH['saida'] + nome_arquivo + f"\img\{hora}.png")
                        if qtde_arquivos == 0:
                            nome_arquivos_imagens = nome_arquivo + f"\img\{hora}.png"
                        else:
                            nome_arquivos_imagens = ", " + nome_arquivo + f"\img\{hora}.png"
                        qtde_arquivos += 1

                # Verifica se tem vídeo
                if len(post.find_elements(By.CSS_SELECTOR, 'video')) > 0:
                    print(f"Salvando Vídeo do post #{i}")
                    hora = str(datetime.datetime.now().hour) + str(datetime.datetime.now().minute) + str(datetime.datetime.now().second) + str(datetime.datetime.now().microsecond)
                    if not os.path.exists(PATH['saida'] + nome_arquivo + f"\movie"):
                        os.makedirs(PATH['saida'] + nome_arquivo + f"\movie")
                    self.baixar_video(url_postagem, PATH['saida'] + nome_arquivo + f"\movie\{hora}.mp4")
                    if qtde_arquivos == 0:
                        nome_arquivos_video = nome_arquivo + f"\movie\{hora}.mp4"
                    else:
                        nome_arquivos_video = ", " + nome_arquivo + f"\movie\{hora}.mp4"
                    qtde_arquivos += 1
                
                if qtde_arquivos > 0:
                    nome_arquivos = nome_arquivos_imagens + nome_arquivos_video
                else:
                    nome_arquivos = ""

                # Salva os dados na planilha
                print(f"Gravando dados do post #{i}")
                self.adicionar_dados_excel(PATH['saida'] + nome_arquivo + f"\dados.xlsx",
                                           data_postagem, 
                                           url_postagem, 
                                           texto_postagem, 
                                           nome_arquivos
                                           )

                # Adiciona post a lista de posts processados
                processados.append(post_id)

            # Parar
            if parar == 1:
                break
            
            # Verifica se todos os posts já foram processados
            if qtde_processados == len(posts):
                break

            # Rola até o último elemento do loop
            ultimo_elemento = posts[-1]
            self.navegador.rolar_para_elemento(ultimo_elemento)
        
        print("Salvando dados no Excel.")
        sleep(1)
        print("Scraping realizado com sucesso!")
        self.navegador.encerrar_navegador()

#----------------------------------------------------------------------------------------

    def carregar(self):
        chars = "/—\|" 
        for _ in range(2):
            for char in chars:
                print(f"\033[34m\rProcessando {char}\033[0m", end="")
                sleep(0.1)

    def carregar_iniciando(self):
        chars = ['.', '..', '...']
        for _ in range(2):
            for char in chars:
                print(f"\033[34m\rIniciando Raspagem {char}\033[0m", end="")
                sleep(0.3)

    def e_url(self, texto=str):
        """Verifica se a string é um Link de URL."""

        if "http://" in texto or "www." in texto or "https://" in texto:
            return True
        else:
            return False

    def login(self, usuario, senha):
        try:
            self.navegador.navegar(LINKS['login'])
            self.navegador.esperar('XPATH', XPATHS['input_login'])
            input_login = self.navegador.obter_elemento('XPATH', XPATHS['input_login'])
            self.navegador.clicar_elemento(input_login)
            self.navegador.enviar_teclas(usuario, espera_entre_as_teclas=0.2)
            self.navegador.enviar_teclas(Keys.ENTER)

            self.navegador.esperar('XPATH', XPATHS['input_password'])
            input_password = self.navegador.obter_elemento('XPATH', XPATHS['input_password'])
            self.navegador.clicar_elemento(input_password)
            self.navegador.enviar_teclas(senha, espera_entre_as_teclas=0.3)
            self.navegador.enviar_teclas(Keys.ENTER)
        except:
            input("Erro ao realizar login! Porfavor, realize o Login manualmente e quando estiver pronto pressione ENTER!")

    def baixar_video(self, link, arquivo):
        janela = self.navegador.id_janela_atual()
        self.navegador.abrir_nova_janela()
        self.navegador.navegar(LINKS['baixar_video'])
        self.navegador.esperar(valor='main_page_text')
        
        input_texto = self.navegador.obter_elemento(valor='main_page_text')
        self.navegador.clicar_elemento(input_texto)
        self.navegador.enviar_teclas(link)
        
        btn_download = self.navegador.obter_elemento(valor='submit')
        self.navegador.clicar_elemento(btn_download)

        while self.navegador.url_atual() == LINKS['baixar_video']:
            pass
        
        self.navegador.esperar('XPATH', XPATHS['videos_download'])
        div_resultados = self.navegador.obter_elemento('XPATH', XPATHS['videos_download'])
        try: 
            a = div_resultados.find_element(By.XPATH, "./a[contains(text(), '540')]")
        except:
            try:
                a = div_resultados.find_element(By.XPATH, "./a[contains(text(), '320')]")
            except:
                try:
                    a = div_resultados.find_element(By.XPATH, "./a[contains(text(), '720')]")
                except:
                    print("Erro ao obter url do vídeo para download.")
        
        url = a.get_attribute("href")
        self.baixar_conteudo(url, arquivo)

        self.navegador.fechar_janela()
        self.navegador.mudar_janela(janela)

    def baixar_conteudo(self, link, nome_arquivo):
        try:
            response = requests.get(link)
            if response.status_code == 200:
                with open(nome_arquivo, 'wb') as arquivo:
                    arquivo.write(response.content)
            else:
                print(f'Erro. Status code: {response.status_code}')
        except Exception as e:
            print(f'Ocorreu um erro: {str(e)}')

    def converte_nome_mes(self, nome_mes):
        meses = {
            'jan': 1,
            'feb': 2,
            'mar': 3,
            'apr': 4,
            'may': 5,
            'jun': 6,
            'jul': 7,
            'aug': 8,
            'sep': 9,
            'oct': 10,
            'nov': 11,
            'dec': 12
        }
        
        # Converte o nome do mês para minúsculas e verifica se está no dicionário
        nome_mes = nome_mes.lower()
        if nome_mes in meses:
            return meses[nome_mes]
        else:
            return None
    
    def adicionar_dados_excel(self, nome_planilha, data, url, texto, nome_arquivos):
        if not os.path.exists(nome_planilha):
            wb = Workbook()
            ws = wb.active
            ws.append(["Data da Postagem", "URL da Postagem", "Texto da Postagem", "Nome dos Arquivos"])
        else:
            wb = load_workbook(nome_planilha)
            ws = wb.active
        
        ws.append([data, url, texto, nome_arquivos])
        wb.save(nome_planilha)
